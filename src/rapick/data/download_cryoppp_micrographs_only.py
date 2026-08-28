#!/usr/bin/env python3
"""
Download ONLY the `.mrc` micrographs for the 34 CryoPPP EMPIAR datasets.

This is a focused re-implementation of the official CryoPPP downloader
(download_micrographs_motion_correction_from_EMPIAR.py). Unlike the official
script it:
  * downloads ONLY the `Micrographs` entries whose path ends in `.mrc`
    (NOT .mrcs / .mrc.gz / .tif / .tiff / .eer, and never Gain / particles /
     ground-truth / star / csv / jpg / png / tar.gz),
  * uses `curl --continue-at -` so interrupted transfers resume,
  * writes to <output-root>/<EMPIAR_ID>/micrographs/*.mrc only,
  * keeps all logs / manifests / failure lists under cryoppp_tools.

URL structure (identical to the official implementation):
  https://ftp.ebi.ac.uk/empiar/world_availability/<Data_Link>/<Micrographs entry>
"""

import argparse
import csv
import os
import shutil
import subprocess
import sys
import time
from concurrent.futures import ThreadPoolExecutor, as_completed
from datetime import datetime
from urllib.parse import quote

import openpyxl

# ----------------------------------------------------------------------------
# Configuration
# ----------------------------------------------------------------------------
BASE_URL = "https://ftp.ebi.ac.uk/empiar/world_availability"

# The 34 EMPIAR IDs that MUST be present in the catalogue.
TARGET_IDS = [
    "10389", "10081", "10289", "11057", "10444", "10576", "10816", "10526", "11051",
    "10760", "11183", "10671", "10291", "10669", "10077", "10061", "10028", "10096",
    "10737", "10387", "10532", "10240", "10005", "10017", "10075", "10184", "10059",
    "10406", "10590", "10093", "10345", "11056", "10852", "10947",
]

# These are set from --data-root at the start of main() via _init_paths().
TOOLS_ROOT = DEFAULT_CATALOG = DEST_ROOT_DEFAULT = None
FAIL_LOG = SKIP_LOG = COLLISION_LOG = MANIFEST = DRYRUN_REPORT = None


def _init_paths(data_root):
    """Populate the module-level run-artifact paths from a data root.
    <root>/cryoppp/ holds data; <root>/cryoppp_tools/ holds logs/manifests."""
    global TOOLS_ROOT, DEFAULT_CATALOG, DEST_ROOT_DEFAULT
    global FAIL_LOG, SKIP_LOG, COLLISION_LOG, MANIFEST, DRYRUN_REPORT
    data_root = os.path.abspath(os.path.expanduser(str(data_root)))
    TOOLS_ROOT = os.path.join(data_root, "cryoppp_tools")
    DEST_ROOT_DEFAULT = os.path.join(data_root, "cryoppp")
    DEFAULT_CATALOG = os.path.join(
        TOOLS_ROOT, "cryoppp",
        "download_micrographs_motion_correction_files",
        "micrographs_download_catalogue.xlsx",
    )
    FAIL_LOG = os.path.join(TOOLS_ROOT, "cryoppp_mrc_failures.tsv")
    SKIP_LOG = os.path.join(TOOLS_ROOT, "cryoppp_mrc_skips.tsv")
    COLLISION_LOG = os.path.join(TOOLS_ROOT, "cryoppp_mrc_collisions.tsv")
    MANIFEST = os.path.join(TOOLS_ROOT, "cryoppp_mrc_manifest.tsv")
    DRYRUN_REPORT = os.path.join(TOOLS_ROOT, "cryoppp_mrc_dryrun_report.txt")
    os.makedirs(TOOLS_ROOT, exist_ok=True)


def ts():
    return datetime.now().strftime("%Y-%m-%d %H:%M:%S")


def url_join(*parts):
    """Join URL parts, collapsing stray slashes and percent-encoding path
    segments (e.g. spaces -> %20) so filenames with spaces (EMPIAR 10077) work.
    `:` and `/` are kept safe so the scheme/host in BASE_URL survive intact."""
    cleaned = [str(p).strip("/") for p in parts if p is not None and str(p).strip("/") != ""]
    return "/".join(quote(c, safe="/:") for c in cleaned)


def load_catalog(catalog_path):
    """Return (data_links, mrc_by_id, skip_reasons, collisions).

    data_links: {id: Data_Link}
    mrc_by_id:  {id: [Micrographs entries ending in .mrc]}
    skip_reasons: {id: reason}  (ids with 0 .mrc entries)
    collisions: {id: {basename: [entries]}}  (basename appears >1 time)
    """
    wb = openpyxl.load_workbook(catalog_path, read_only=True, data_only=True)

    # --- EMPIAR-IDs sheet ---
    if "EMPIAR-IDs" not in wb.sheetnames:
        raise SystemExit("FATAL: 'EMPIAR-IDs' sheet not found in catalogue.")
    id_ws = wb["EMPIAR-IDs"]
    id_rows = list(id_ws.iter_rows(values_only=True))
    header = list(id_rows[0])
    col_id = header.index("EMPIAR_ID")
    col_link = header.index("Data_Link")

    data_links = {}
    catalog_ids = []
    for r in id_rows[1:]:
        if r[col_id] is None:
            continue
        sid = str(r[col_id]).strip()
        catalog_ids.append(sid)
        data_links[sid] = r[col_link]

    # --- Verify target set vs catalogue ---
    catalog_set = set(catalog_ids)
    target_set = set(TARGET_IDS)
    missing = sorted(target_set - catalog_set)
    extra = sorted(catalog_set - target_set)
    if missing or extra:
        print("FATAL: target ID set does not match catalogue EMPIAR-IDs sheet.", file=sys.stderr)
        print("  missing from catalogue:", missing, file=sys.stderr)
        print("  extra in catalogue   :", extra, file=sys.stderr)
        raise SystemExit(2)
    if len(target_set) != 34:
        raise SystemExit(f"FATAL: expected 34 target IDs, got {len(target_set)}")

    # --- Per-ID Micrographs columns ---
    mrc_by_id = {}
    skip_reasons = {}
    collisions = {}
    for sid in TARGET_IDS:
        if sid not in wb.sheetnames:
            skip_reasons[sid] = "no per-ID sheet in catalogue"
            mrc_by_id[sid] = []
            continue
        ws = wb[sid]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            skip_reasons[sid] = "empty sheet"
            mrc_by_id[sid] = []
            continue
        h = list(rows[0])
        if "Micrographs" not in h:
            skip_reasons[sid] = "no 'Micrographs' column"
            mrc_by_id[sid] = []
            continue
        mi = h.index("Micrographs")
        entries = [row[mi] for row in rows[1:] if row[mi] not in (None, "")]
        # ONLY paths ending in .mrc (excludes .mrcs, .mrc.gz, .tif, .tiff, .eer)
        mrc = [str(e) for e in entries if str(e).lower().endswith(".mrc")]
        mrc_by_id[sid] = mrc
        if not mrc:
            n_other = len(entries)
            skip_reasons[sid] = f"0 .mrc entries (Micrographs rows={n_other}, all non-.mrc)"
            continue
        # basename collision detection
        seen = {}
        for e in mrc:
            bn = os.path.basename(str(e))
            seen.setdefault(bn, []).append(e)
        coll = {bn: lst for bn, lst in seen.items() if len(lst) > 1}
        if coll:
            collisions[sid] = coll

    wb.close()
    return data_links, mrc_by_id, skip_reasons, collisions


def curl_content_length(url, timeout=60):
    """Return Content-Length in bytes via a HEAD request, or None."""
    try:
        out = subprocess.run(
            ["curl", "-sIL", "--max-time", str(timeout), url],
            capture_output=True, text=True,
        )
        cl = None
        for line in out.stdout.splitlines():
            if line.lower().startswith("content-length:"):
                cl = int(line.split(":", 1)[1].strip())
        return cl
    except Exception:
        return None


def human(n):
    if n is None:
        return "?"
    for unit in ["B", "KB", "MB", "GB", "TB", "PB"]:
        if abs(n) < 1024.0:
            return f"{n:.2f}{unit}"
        n /= 1024.0
    return f"{n:.2f}EB"


def estimate_sizes(data_links, mrc_by_id, ids, sample_per_id=3):
    """Sample up to `sample_per_id` files per ID via HEAD, extrapolate per-ID size.

    Returns (per_id_estimate{id:bytes}, total_bytes, detail{id:(samples,avg)}).
    """
    per_id = {}
    detail = {}
    total = 0
    for sid in ids:
        mrc = mrc_by_id[sid]
        if not mrc:
            per_id[sid] = 0
            detail[sid] = (0, 0)
            continue
        link = data_links[sid]
        sample = mrc[:sample_per_id]
        sizes = []
        for entry in sample:
            url = url_join(BASE_URL, link, entry)
            cl = curl_content_length(url)
            if cl:
                sizes.append(cl)
        if sizes:
            avg = sum(sizes) / len(sizes)
        else:
            avg = 0
        est = int(avg * len(mrc))
        per_id[sid] = est
        detail[sid] = (len(sizes), avg)
        total += est
    return per_id, total, detail


def download_one(sid, entry, data_link, out_dir, max_retries):
    """Download a single .mrc with curl --continue-at -. Returns (ok, msg, code)."""
    bn = os.path.basename(str(entry))
    final = os.path.join(out_dir, bn)
    part = final + ".part"
    url = url_join(BASE_URL, data_link, entry)

    # skip if already complete
    if os.path.exists(final) and os.path.getsize(final) > 0:
        return (True, "skip-exists", 0)

    code = -1
    for attempt in range(1, max_retries + 1):
        proc = subprocess.run(
            ["curl", "-fSL", "--continue-at", "-",
             "--retry", "0", "--connect-timeout", "30",
             "-o", part, url],
            capture_output=True, text=True,
        )
        code = proc.returncode
        if code == 0 and os.path.exists(part) and os.path.getsize(part) > 0:
            os.replace(part, final)
            return (True, f"ok(attempt {attempt})", 0)
        # curl exit 33 => server doesn't support resume; restart fresh
        if code == 33 and os.path.exists(part):
            try:
                os.remove(part)
            except OSError:
                pass
        time.sleep(min(5 * attempt, 30))
    return (False, f"curl exit {code} after {max_retries} retries", code)


def append_tsv(path, row, header=None):
    new = not os.path.exists(path)
    with open(path, "a", newline="") as f:
        w = csv.writer(f, delimiter="\t")
        if new and header:
            w.writerow(header)
        w.writerow(row)


def main():
    ap = argparse.ArgumentParser(description="Download CryoPPP .mrc micrographs only.")
    ap.add_argument("--data-root", required=True,
                    help="REQUIRED. Data root; data lands under <root>/cryoppp/, "
                         "run artifacts under <root>/cryoppp_tools/.")
    ap.add_argument("--all", action="store_true", help="Process all 34 target EMPIAR IDs.")
    ap.add_argument("--ids", nargs="+", metavar="ID",
                    help="Process only these EMPIAR IDs instead of all 34. The catalogue is "
                         "still validated in full, so a narrowed run cannot mask a bad "
                         "catalogue; only the download set shrinks.")
    ap.add_argument("--output-root", default=None,
                    help="Override the output dir (default <data-root>/cryoppp).")
    ap.add_argument("--workers", type=int, default=1, help="Concurrent downloads (default 1).")
    ap.add_argument("--dry-run", action="store_true", help="Plan + capacity check only.")
    ap.add_argument("--max-retries", type=int, default=5, help="Per-file curl retries.")
    ap.add_argument("--catalog", default=None,
                    help="Override the xlsx catalogue path (default <data-root>/cryoppp_tools/"
                         "cryoppp/download_micrographs_motion_correction_files/"
                         "micrographs_download_catalogue.xlsx).")
    ap.add_argument("--sample-per-id", type=int, default=3,
                    help="Files sampled per ID for size estimate in dry-run.")
    args = ap.parse_args()

    if not args.all and not args.ids:
        print("Pass --all (the 34 CryoPPP target IDs) or --ids <ID...>.", file=sys.stderr)
        sys.exit(2)

    selected_ids = TARGET_IDS
    if args.ids:
        unknown = sorted(set(args.ids) - set(TARGET_IDS))
        if unknown:
            print(f"Not CryoPPP target IDs: {unknown}", file=sys.stderr)
            sys.exit(2)
        # Keep TARGET_IDS order so reports read the same however --ids was typed.
        selected_ids = [sid for sid in TARGET_IDS if sid in set(args.ids)]

    _init_paths(args.data_root)   # sets TOOLS_ROOT / logs / DEST_ROOT_DEFAULT / DEFAULT_CATALOG
    args.catalog = args.catalog or DEFAULT_CATALOG
    out_root = os.path.expanduser(args.output_root) if args.output_root else DEST_ROOT_DEFAULT
    print(f"[{ts()}] catalogue: {args.catalog}")
    if not os.path.exists(args.catalog):
        raise SystemExit(f"FATAL: catalogue not found: {args.catalog}")

    data_links, mrc_by_id, skip_reasons, collisions = load_catalog(args.catalog)

    n_ids = len(TARGET_IDS)
    total_mrc = sum(len(v) for v in mrc_by_id.values())
    print(f"[{ts()}] EMPIAR IDs detected in catalogue: {n_ids}")
    print(f"[{ts()}] total .mrc target files: {total_mrc}")

    # Per-ID summary
    print("\n  ID        .mrc   note")
    for sid in selected_ids:
        note = ""
        if sid in skip_reasons:
            note = "SKIP: " + skip_reasons[sid]
        if sid in collisions:
            note = (note + " | " if note else "") + f"COLLISION: {collisions[sid]}"
        print(f"  {sid:8} {len(mrc_by_id[sid]):5}   {note}")

    # Record skips
    if skip_reasons:
        for sid, reason in skip_reasons.items():
            append_tsv(SKIP_LOG, [ts(), sid, reason],
                       header=["timestamp", "empiar_id", "reason"])
    # Record collisions -> these IDs are failures, do not download
    if collisions:
        for sid, coll in collisions.items():
            for bn, lst in coll.items():
                append_tsv(COLLISION_LOG, [ts(), sid, bn, ";".join(map(str, lst))],
                           header=["timestamp", "empiar_id", "basename", "entries"])

    collision_ids = set(collisions.keys())

    if args.dry_run:
        print(f"\n[{ts()}] estimating sizes (HEAD sampling up to {args.sample_per_id}/ID)...")
        per_id, total_bytes, detail = estimate_sizes(
            data_links, mrc_by_id, selected_ids, args.sample_per_id)
        usage = shutil.disk_usage(os.path.dirname(out_root) if not os.path.exists(out_root) else out_root)
        free = usage.free
        needed = total_bytes
        margin_ok = free >= needed * 1.10
        shortfall = max(0, int(needed * 1.10) - free)

        lines = []
        lines.append(f"CryoPPP .mrc dry-run report  {ts()}")
        lines.append(f"catalogue: {args.catalog}")
        lines.append(f"output-root: {out_root}")
        lines.append(f"EMPIAR IDs in catalogue: {n_ids} (expected 34)")
        lines.append(f"total .mrc target files: {total_mrc}")
        lines.append("")
        lines.append(f"{'ID':8} {'.mrc':>6} {'samples':>8} {'avg/file':>12} {'est size':>12}")
        for sid in selected_ids:
            ns, avg = detail[sid]
            lines.append(f"{sid:8} {len(mrc_by_id[sid]):6} {ns:8} {human(avg):>12} {human(per_id[sid]):>12}")
        lines.append("")
        lines.append(f"estimated total needed : {human(needed)} ({needed} bytes)")
        lines.append(f"free on data root      : {human(free)} ({free} bytes)")
        lines.append(f"needed + 10% margin    : {human(int(needed*1.10))}")
        lines.append(f"10% margin satisfied   : {'YES' if margin_ok else 'NO'}")
        if not margin_ok:
            lines.append(f"shortfall              : {human(shortfall)} ({shortfall} bytes)")
        if skip_reasons:
            lines.append("")
            lines.append("SKIPPED IDs (0 .mrc):")
            for sid, r in skip_reasons.items():
                lines.append(f"  {sid}: {r}")
        if collisions:
            lines.append("")
            lines.append("BASENAME COLLISIONS (treated as failures, not downloaded):")
            for sid, c in collisions.items():
                lines.append(f"  {sid}: {c}")

        report = "\n".join(lines)
        print("\n" + report)
        with open(DRYRUN_REPORT, "w") as f:
            f.write(report + "\n")
        print(f"\n[{ts()}] dry-run report written to {DRYRUN_REPORT}")

        if not margin_ok:
            print(f"\n[{ts()}] INSUFFICIENT SPACE — not starting download.", file=sys.stderr)
            print(f"  needed (+10%): {human(int(needed*1.10))}  free: {human(free)}  short: {human(shortfall)}",
                  file=sys.stderr)
            sys.exit(3)
        print(f"\n[{ts()}] capacity OK (>=10% margin). Dry-run complete.")
        return

    # ---------------- Real download ----------------
    print(f"\n[{ts()}] starting REAL download. workers={args.workers} max-retries={args.max_retries}")
    # Build work list, skipping 0-mrc IDs and collision IDs (failures)
    tasks = []
    for sid in selected_ids:
        if not mrc_by_id[sid]:
            continue
        if sid in collision_ids:
            print(f"[{ts()}] {sid}: SKIPPING download due to basename collision (marked failed).")
            for entry in mrc_by_id[sid]:
                append_tsv(FAIL_LOG,
                           [ts(), sid, os.path.basename(str(entry)),
                            url_join(BASE_URL, data_links[sid], entry),
                            "basename-collision", "skipped"],
                           header=["timestamp", "empiar_id", "filename", "url", "exit_code", "note"])
            continue
        out_dir = os.path.join(out_root, sid, "micrographs")
        os.makedirs(out_dir, exist_ok=True)
        for entry in mrc_by_id[sid]:
            tasks.append((sid, entry, data_links[sid], out_dir))

    print(f"[{ts()}] queued {len(tasks)} file downloads across "
          f"{len(set(t[0] for t in tasks))} IDs.")

    ok_count = 0
    fail_count = 0
    skip_count = 0

    def run(task):
        sid, entry, link, out_dir = task
        return (sid, entry, link) + download_one(sid, entry, link, out_dir, args.max_retries)

    if args.workers <= 1:
        results = (run(t) for t in tasks)
        for i, (sid, entry, link, ok, msg, code) in enumerate(results, 1):
            bn = os.path.basename(str(entry))
            if ok and msg == "skip-exists":
                skip_count += 1
            elif ok:
                ok_count += 1
                append_tsv(MANIFEST, [ts(), sid, bn, url_join(BASE_URL, link, entry)],
                           header=["timestamp", "empiar_id", "filename", "url"])
            else:
                fail_count += 1
                append_tsv(FAIL_LOG, [ts(), sid, bn, url_join(BASE_URL, link, entry), code, msg],
                           header=["timestamp", "empiar_id", "filename", "url", "exit_code", "note"])
            if i % 50 == 0 or i == len(tasks):
                print(f"[{ts()}] progress {i}/{len(tasks)} ok={ok_count} skip={skip_count} fail={fail_count}")
    else:
        with ThreadPoolExecutor(max_workers=args.workers) as ex:
            futs = {ex.submit(run, t): t for t in tasks}
            done = 0
            for fut in as_completed(futs):
                done += 1
                sid, entry, link, ok, msg, code = fut.result()
                bn = os.path.basename(str(entry))
                if ok and msg == "skip-exists":
                    skip_count += 1
                elif ok:
                    ok_count += 1
                    append_tsv(MANIFEST, [ts(), sid, bn, url_join(BASE_URL, link, entry)],
                               header=["timestamp", "empiar_id", "filename", "url"])
                else:
                    fail_count += 1
                    append_tsv(FAIL_LOG, [ts(), sid, bn, url_join(BASE_URL, link, entry), code, msg],
                               header=["timestamp", "empiar_id", "filename", "url", "exit_code", "note"])
                if done % 50 == 0 or done == len(tasks):
                    print(f"[{ts()}] progress {done}/{len(tasks)} ok={ok_count} skip={skip_count} fail={fail_count}")

    print(f"\n[{ts()}] DONE. ok={ok_count} skip-exists={skip_count} fail={fail_count}")
    print(f"  manifest : {MANIFEST}")
    print(f"  failures : {FAIL_LOG}")
    print(f"  skips    : {SKIP_LOG}")
    print(f"  collisions: {COLLISION_LOG}")


if __name__ == "__main__":
    main()
